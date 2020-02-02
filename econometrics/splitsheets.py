from openpyxl import load_workbook, Workbook

wb = load_workbook(filename='G:\\code\\All operation\\table3b.xlsx')

for sheet in wb.worksheets:
    new_wb = Workbook()
    ws = new_wb.active
    for row_data in sheet.iter_rows():
        for row_cell in row_data:
            ws[row_cell.coordinate].value = row_cell.value

    new_wb.save('{0}.xlsx'.format(sheet.title))
